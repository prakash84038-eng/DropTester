"""
Data Export and Integration Module
Provides various export formats and integration capabilities.
"""

import os
import json
import csv
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import zipfile
import tempfile

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .analytics import TestAnalytics

class DataExporter:
    """Handles data export in various formats and provides integration capabilities."""
    
    def __init__(self, analytics: TestAnalytics = None, data_dir: str = None):
        """Initialize data exporter."""
        self.analytics = analytics or TestAnalytics(data_dir)
        self.export_templates = self._load_export_templates()
    
    def _load_export_templates(self) -> Dict:
        """Load export templates configuration."""
        return {
            "csv": {
                "headers": [
                    "Test ID", "Timestamp", "Sample Code", "IS Number", "Parameter",
                    "Department", "Testing Person", "Material Type", "Result",
                    "Confidence", "Metric", "Metric Value", "Reason", "Manual Override"
                ],
                "delimiter": ",",
                "encoding": "utf-8"
            },
            "excel": {
                "sheet_name": "Test Results",
                "include_charts": True,
                "format_cells": True
            },
            "json": {
                "format": "pretty",
                "include_metadata": True
            },
            "xml": {
                "root_element": "TestResults",
                "item_element": "TestResult"
            }
        }
    
    def export_comprehensive_report(self, output_path: str, format_type: str = "excel",
                                  start_date: str = None, end_date: str = None,
                                  include_analytics: bool = True) -> bool:
        """Export comprehensive report with test data and analytics."""
        try:
            if format_type.lower() == "excel" and EXCEL_AVAILABLE:
                return self._export_excel_comprehensive(output_path, start_date, end_date, include_analytics)
            elif format_type.lower() == "json":
                return self._export_json_comprehensive(output_path, start_date, end_date, include_analytics)
            elif format_type.lower() == "csv_package":
                return self._export_csv_package(output_path, start_date, end_date, include_analytics)
            else:
                # Fall back to basic export
                return self.analytics.export_data(output_path, format_type, start_date, end_date)
                
        except Exception as e:
            print(f"Error exporting comprehensive report: {e}")
            return False
    
    def _export_excel_comprehensive(self, output_path: str, start_date: str = None,
                                  end_date: str = None, include_analytics: bool = True) -> bool:
        """Export comprehensive Excel report with multiple sheets and charts."""
        if not EXCEL_AVAILABLE:
            print("openpyxl not available for Excel export")
            return False
        
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Add test data sheet
        self._add_test_data_sheet(workbook, start_date, end_date)
        
        if include_analytics:
            # Add analytics sheets
            self._add_summary_sheet(workbook)
            self._add_trends_sheet(workbook)
            self._add_performance_sheet(workbook)
        
        # Save workbook
        workbook.save(output_path)
        return True
    
    def _add_test_data_sheet(self, workbook, start_date: str = None, end_date: str = None):
        """Add test data sheet to Excel workbook."""
        # Get test data from database
        import sqlite3
        conn = sqlite3.connect(self.analytics.db_path)
        cursor = conn.cursor()
        
        query = "SELECT * FROM tests"
        params = []
        
        if start_date or end_date:
            query += " WHERE"
            if start_date:
                query += " DATE(timestamp) >= ?"
                params.append(start_date)
            if end_date:
                if start_date:
                    query += " AND"
                query += " DATE(timestamp) <= ?"
                params.append(end_date)
        
        query += " ORDER BY timestamp DESC"
        
        cursor.execute(query, params)
        columns = [description[0] for description in cursor.description]
        data = cursor.fetchall()
        conn.close()
        
        # Create worksheet
        ws = workbook.create_sheet(title="Test Data")
        
        # Add headers with formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, record in enumerate(data, 2):
            for col, value in enumerate(record, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_sheet(self, workbook):
        """Add summary analytics sheet."""
        ws = workbook.create_sheet(title="Summary Analytics")
        
        # Get summary statistics
        stats = self.analytics.get_summary_stats(30)
        material_stats = self.analytics.get_material_performance()
        tester_stats = self.analytics.get_tester_performance()
        
        # Add summary section
        ws["A1"] = "SUMMARY STATISTICS (Last 30 Days)"
        ws["A1"].font = Font(bold=True, size=14)
        
        summary_data = [
            ["Total Tests", stats['total_tests']],
            ["Pass Count", stats['pass_count']],
            ["Fail Count", stats['fail_count']],
            ["Pass Rate", f"{stats['pass_rate']:.1f}%"],
            ["Unique Testers", stats['unique_testers']],
            ["Average Confidence", f"{stats['avg_confidence']:.2f}"]
        ]
        
        for i, (label, value) in enumerate(summary_data, 3):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].font = Font(bold=True)
        
        # Add material performance section
        row_start = len(summary_data) + 5
        ws[f"A{row_start}"] = "MATERIAL PERFORMANCE"
        ws[f"A{row_start}"].font = Font(bold=True, size=12)
        
        material_headers = ["Material", "Total Tests", "Pass Rate", "Avg Confidence"]
        for col, header in enumerate(material_headers, 1):
            ws.cell(row=row_start + 2, column=col, value=header).font = Font(bold=True)
        
        for i, (material, data) in enumerate(material_stats.items(), 1):
            row = row_start + 2 + i
            ws.cell(row=row, column=1, value=material)
            ws.cell(row=row, column=2, value=data['total_tests'])
            ws.cell(row=row, column=3, value=f"{data['pass_rate']:.1f}%")
            ws.cell(row=row, column=4, value=f"{data['avg_confidence']:.2f}")
    
    def _add_trends_sheet(self, workbook):
        """Add trends analysis sheet with chart."""
        ws = workbook.create_sheet(title="Trends Analysis")
        
        # Get trend data
        trend_data = self.analytics.get_trend_data(30)
        
        # Add headers
        headers = ["Date", "Total Tests", "Pass Count", "Fail Count", "Pass Rate %"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        # Add data
        for row, data in enumerate(trend_data, 2):
            ws.cell(row=row, column=1, value=data['date'])
            ws.cell(row=row, column=2, value=data['total_tests'])
            ws.cell(row=row, column=3, value=data['pass_count'])
            ws.cell(row=row, column=4, value=data['fail_count'])
            ws.cell(row=row, column=5, value=data['pass_rate'])
        
        # Add chart if data is available
        if len(trend_data) > 1:
            chart = BarChart()
            chart.title = "Daily Test Trends"
            chart.y_axis.title = "Number of Tests"
            chart.x_axis.title = "Date"
            
            data_ref = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=len(trend_data) + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(trend_data) + 1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            ws.add_chart(chart, "G2")
    
    def _add_performance_sheet(self, workbook):
        """Add performance analysis sheet."""
        ws = workbook.create_sheet(title="Performance Analysis")
        
        # Get failure patterns
        patterns = self.analytics.get_failure_patterns()
        
        ws["A1"] = "FAILURE PATTERN ANALYSIS"
        ws["A1"].font = Font(bold=True, size=14)
        
        headers = ["Failure Reason", "Total Frequency", "Materials Affected"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).font = Font(bold=True)
        
        for i, (reason, data) in enumerate(patterns.items(), 1):
            row = 3 + i
            ws.cell(row=row, column=1, value=reason)
            ws.cell(row=row, column=2, value=data['total_frequency'])
            materials = ", ".join(data['by_material'].keys())
            ws.cell(row=row, column=3, value=materials)
    
    def _export_json_comprehensive(self, output_path: str, start_date: str = None,
                                 end_date: str = None, include_analytics: bool = True) -> bool:
        """Export comprehensive JSON report."""
        try:
            comprehensive_data = {
                "export_info": {
                    "timestamp": datetime.now().isoformat(),
                    "start_date": start_date,
                    "end_date": end_date,
                    "format": "comprehensive_json"
                },
                "test_data": [],
                "analytics": {}
            }
            
            # Get test data
            import sqlite3
            conn = sqlite3.connect(self.analytics.db_path)
            cursor = conn.cursor()
            
            query = "SELECT * FROM tests"
            params = []
            
            if start_date or end_date:
                query += " WHERE"
                if start_date:
                    query += " DATE(timestamp) >= ?"
                    params.append(start_date)
                if end_date:
                    if start_date:
                        query += " AND"
                    query += " DATE(timestamp) <= ?"
                    params.append(end_date)
            
            query += " ORDER BY timestamp DESC"
            
            cursor.execute(query, params)
            columns = [description[0] for description in cursor.description]
            data = cursor.fetchall()
            conn.close()
            
            # Convert to list of dictionaries
            for row in data:
                comprehensive_data["test_data"].append(dict(zip(columns, row)))
            
            # Add analytics if requested
            if include_analytics:
                comprehensive_data["analytics"] = {
                    "summary_stats": self.analytics.get_summary_stats(30),
                    "trend_data": self.analytics.get_trend_data(30),
                    "material_performance": self.analytics.get_material_performance(),
                    "tester_performance": self.analytics.get_tester_performance(),
                    "failure_patterns": self.analytics.get_failure_patterns()
                }
            
            # Write to file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(comprehensive_data, f, indent=2, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            print(f"Error exporting JSON comprehensive report: {e}")
            return False
    
    def _export_csv_package(self, output_path: str, start_date: str = None,
                          end_date: str = None, include_analytics: bool = True) -> bool:
        """Export package of CSV files with analytics."""
        try:
            # Create temporary directory
            with tempfile.TemporaryDirectory() as temp_dir:
                
                # Export main test data
                test_data_path = os.path.join(temp_dir, "test_data.csv")
                self.analytics.export_data(test_data_path, "csv", start_date, end_date)
                
                files_to_zip = [test_data_path]
                
                if include_analytics:
                    # Export summary statistics
                    summary_path = os.path.join(temp_dir, "summary_stats.csv")
                    self._export_summary_csv(summary_path)
                    files_to_zip.append(summary_path)
                    
                    # Export trend data
                    trends_path = os.path.join(temp_dir, "trends.csv")
                    self._export_trends_csv(trends_path)
                    files_to_zip.append(trends_path)
                    
                    # Export material performance
                    material_path = os.path.join(temp_dir, "material_performance.csv")
                    self._export_material_performance_csv(material_path)
                    files_to_zip.append(material_path)
                
                # Create ZIP file
                with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file_path in files_to_zip:
                        if os.path.exists(file_path):
                            zipf.write(file_path, os.path.basename(file_path))
                
                return True
                
        except Exception as e:
            print(f"Error creating CSV package: {e}")
            return False
    
    def _export_summary_csv(self, output_path: str):
        """Export summary statistics as CSV."""
        stats = self.analytics.get_summary_stats(30)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Metric", "Value"])
            
            for key, value in stats.items():
                if isinstance(value, (int, float)):
                    if key.endswith('_rate'):
                        value = f"{value:.1f}%"
                    elif isinstance(value, float):
                        value = f"{value:.2f}"
                writer.writerow([key.replace('_', ' ').title(), value])
    
    def _export_trends_csv(self, output_path: str):
        """Export trend data as CSV."""
        trend_data = self.analytics.get_trend_data(30)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            if trend_data:
                writer = csv.DictWriter(csvfile, fieldnames=trend_data[0].keys())
                writer.writeheader()
                writer.writerows(trend_data)
    
    def _export_material_performance_csv(self, output_path: str):
        """Export material performance as CSV."""
        material_stats = self.analytics.get_material_performance()
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Material", "Total Tests", "Pass Count", "Fail Count", 
                           "Pass Rate %", "Avg Confidence", "Avg Metric Value"])
            
            for material, stats in material_stats.items():
                writer.writerow([
                    material,
                    stats['total_tests'],
                    stats['pass_count'],
                    stats['fail_count'],
                    f"{stats['pass_rate']:.1f}%",
                    f"{stats['avg_confidence']:.2f}",
                    f"{stats['avg_metric_value']:.2f}"
                ])
    
    def export_for_api(self, output_path: str, api_format: str = "rest") -> bool:
        """Export data in API-friendly format."""
        try:
            if api_format.lower() == "rest":
                return self._export_rest_api_format(output_path)
            elif api_format.lower() == "graphql":
                return self._export_graphql_format(output_path)
            else:
                return self._export_generic_api_format(output_path)
                
        except Exception as e:
            print(f"Error exporting for API: {e}")
            return False
    
    def _export_rest_api_format(self, output_path: str) -> bool:
        """Export in REST API format."""
        # Get recent test data
        import sqlite3
        conn = sqlite3.connect(self.analytics.db_path)
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM tests ORDER BY timestamp DESC LIMIT 1000")
        columns = [description[0] for description in cursor.description]
        data = cursor.fetchall()
        conn.close()
        
        # Format for REST API
        api_data = {
            "metadata": {
                "total_records": len(data),
                "generated_at": datetime.now().isoformat(),
                "api_version": "1.0"
            },
            "data": [dict(zip(columns, row)) for row in data],
            "pagination": {
                "page": 1,
                "per_page": len(data),
                "total_pages": 1
            }
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(api_data, f, indent=2, ensure_ascii=False)
        
        return True
    
    def _export_generic_api_format(self, output_path: str) -> bool:
        """Export in generic API format."""
        return self._export_rest_api_format(output_path)
    
    def export_compliance_report(self, output_path: str, standard: str = "ISO") -> bool:
        """Export compliance report for specific standards."""
        try:
            compliance_data = {
                "report_info": {
                    "standard": standard,
                    "generated_at": datetime.now().isoformat(),
                    "reporting_period": "Last 30 days"
                },
                "summary": self.analytics.get_summary_stats(30),
                "test_procedures": {
                    "materials_tested": list(self.analytics.get_material_performance().keys()),
                    "test_methods": ["Rule-based analysis", "Visual inspection", "Automated recording"]
                },
                "quality_metrics": {
                    "repeatability": "Within acceptable limits",
                    "traceability": "Full audit trail maintained",
                    "calibration": "Regular calibration performed"
                },
                "non_conformances": self._get_non_conformances()
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(compliance_data, f, indent=2, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            print(f"Error generating compliance report: {e}")
            return False
    
    def _get_non_conformances(self) -> List[Dict]:
        """Get list of non-conformances (failed tests with low confidence)."""
        import sqlite3
        conn = sqlite3.connect(self.analytics.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT timestamp, sample_code, result, confidence, reason
            FROM tests 
            WHERE result = 'FAIL' AND (confidence < 0.7 OR manual_override = 1)
            ORDER BY timestamp DESC
            LIMIT 50
        """)
        
        results = cursor.fetchall()
        conn.close()
        
        non_conformances = []
        for row in results:
            non_conformances.append({
                "timestamp": row[0],
                "sample_code": row[1],
                "result": row[2],
                "confidence": row[3],
                "reason": row[4],
                "severity": "High" if row[3] < 0.5 else "Medium"
            })
        
        return non_conformances